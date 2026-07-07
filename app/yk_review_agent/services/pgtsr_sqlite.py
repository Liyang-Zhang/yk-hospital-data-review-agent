from __future__ import annotations

import sqlite3
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from yk_review_agent.core.config import settings
from yk_review_agent.services.pgta_sqlite import (
    IMPORT_BATCH_TABLE,
    _connect,
    _normalize_header,
    _normalize_month_bucket_value,
    _sqlite_path_from_url,
    _to_int,
    _to_iso_datetime,
)
from yk_review_agent.services.pgta_detail_dataset import (
    _normalize_qc_conclusion,
    _normalize_result_label,
    _normalize_sample_type,
    _normalize_text,
    _parse_datetime,
    _to_float,
)


PGTSR_TABLE_NAME = "pgtsr_snapshot_raw"


@dataclass(frozen=True)
class PGTSRRecord:
    created_at: datetime
    month_bucket: str
    cycle_id: str
    sample_id: str
    sample_name: str
    sample_type: str
    hospital_name: str
    product_name: str
    result_label: str
    qc_conclusion: str
    cnv_result: str
    cnv_hint: str
    result_detail: str
    patient_age: int | None
    spouse_age: int | None
    patient_karyotype: str
    spouse_karyotype: str
    sr_clinical_type: str
    next_step_screening: str

    @property
    def has_main_cnv_result(self) -> bool:
        return self.cnv_result not in {"", "/"}

    @property
    def is_qc_fail(self) -> bool:
        return self.qc_conclusion == "FAIL" or self.result_label == "质控不合格"

    @property
    def is_euploid(self) -> bool:
        return self.result_label == "未见异常"


PGTSR_SCHEMA: list[tuple[str, str]] = [
    ("id", "INTEGER PRIMARY KEY AUTOINCREMENT"),
    ("import_batch_id", "INTEGER NOT NULL"),
    ("source_file_name", "TEXT NOT NULL"),
    ("source_sheet_name", "TEXT NOT NULL"),
    ("source_row_num", "INTEGER NOT NULL"),
    ("imported_at", "TEXT NOT NULL"),
    ("month_bucket", "TEXT"),
    ("report_review_time", "TEXT"),
    ("hospital_name", "TEXT"),
    ("product_code", "TEXT"),
    ("cycle_id", "TEXT"),
    ("sample_id", "TEXT"),
    ("sample_name", "TEXT"),
    ("sample_type", "TEXT"),
    ("patient_age", "INTEGER"),
    ("spouse_age", "INTEGER"),
    ("patient_karyotype", "TEXT"),
    ("spouse_karyotype", "TEXT"),
    ("data_qc_conclusion", "TEXT"),
    ("cnv_result", "TEXT"),
    ("result_label", "TEXT"),
    ("cnv_hint", "TEXT"),
    ("result_detail", "TEXT"),
    ("sr_clinical_type", "TEXT"),
    ("next_step_screening", "TEXT"),
]


def create_pgtsr_snapshot_schema(database_url: str) -> None:
    with _connect(database_url) as conn:
        conn.execute(
            f"""
            CREATE TABLE IF NOT EXISTS {PGTSR_TABLE_NAME} (
                {", ".join(f"{name} {definition}" for name, definition in PGTSR_SCHEMA)}
            )
            """
        )
        conn.execute(
            f"CREATE INDEX IF NOT EXISTS idx_pgtsr_hospital_month ON {PGTSR_TABLE_NAME}(hospital_name, month_bucket)"
        )
        conn.execute(
            f"CREATE INDEX IF NOT EXISTS idx_pgtsr_cycle ON {PGTSR_TABLE_NAME}(cycle_id)"
        )
        conn.commit()


class PGTSRExcelImporter:
    def __init__(self, database_url: str) -> None:
        self.database_url = database_url
        self.source_file = Path(settings.pgtsr_snapshot_file).resolve()

    def rebuild(self) -> None:
        create_pgtsr_snapshot_schema(self.database_url)
        with _connect(self.database_url) as conn:
            conn.execute(f"DELETE FROM {PGTSR_TABLE_NAME}")
            conn.execute(f"DELETE FROM {IMPORT_BATCH_TABLE} WHERE product_code = 'PGT-SR'")
            conn.commit()
        self._import_workbook()

    def _import_workbook(self) -> None:
        if not self.source_file.exists():
            raise RuntimeError(f"PGTSR source file not found: {self.source_file}")

        started_at = datetime.utcnow().isoformat(sep=" ")
        with _connect(self.database_url) as conn:
            cursor = conn.execute(
                f"""
                INSERT INTO {IMPORT_BATCH_TABLE} (
                    product_code, table_name, source_file_name, source_sheet_name,
                    source_file_size, source_file_mtime, import_started_at, status
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    "PGT-SR",
                    PGTSR_TABLE_NAME,
                    self.source_file.name,
                    "2025年-数据 + 2025年-项目",
                    self.source_file.stat().st_size,
                    datetime.fromtimestamp(self.source_file.stat().st_mtime).isoformat(sep=" "),
                    started_at,
                    "running",
                ),
            )
            batch_id = int(cursor.lastrowid)
            conn.commit()

        raw_count = 0
        loaded_count = 0
        status = "success"
        notes = ""
        try:
            raw_count, loaded_count = self._import_rows(batch_id)
        except Exception as exc:
            status = "failed"
            notes = str(exc)
            raise
        finally:
            with _connect(self.database_url) as conn:
                conn.execute(
                    f"""
                    UPDATE {IMPORT_BATCH_TABLE}
                    SET import_finished_at = ?, row_count_raw = ?, row_count_loaded = ?, status = ?, notes = ?
                    WHERE id = ?
                    """,
                    (
                        datetime.utcnow().isoformat(sep=" "),
                        raw_count,
                        loaded_count,
                        status,
                        notes,
                        batch_id,
                    ),
                )
                conn.commit()

    def _import_rows(self, batch_id: int) -> tuple[int, int]:
        workbook = load_workbook(self.source_file, read_only=True, data_only=True)
        if "2025年-数据" not in workbook.sheetnames:
            raise RuntimeError("PGTSR workbook must contain 2025年-数据 sheet.")

        project_map: dict[str, dict[str, str]] = {}
        if "2025年-项目" in workbook.sheetnames:
            project_sheet = workbook["2025年-项目"]
            project_header = next(project_sheet.iter_rows(min_row=1, max_row=1, values_only=True))
            project_index = {
                _normalize_header(value): position
                for position, value in enumerate(project_header)
                if _normalize_header(value)
            }
            for row in project_sheet.iter_rows(min_row=2, values_only=True):
                if not any(value is not None and str(value).strip() for value in row):
                    continue
                cycle_id = self._raw(project_index, row, "送检单编号")
                if not cycle_id:
                    continue
                project_map[str(cycle_id).strip()] = {
                    "sr_clinical_type": _normalize_text(self._raw(project_index, row, "临床指征")),
                    "next_step_screening": _normalize_text(self._raw(project_index, row, "是否进入下一步易位筛查")),
                }

        data_sheet = workbook["2025年-数据"]
        data_header = next(data_sheet.iter_rows(min_row=1, max_row=1, values_only=True))
        data_index = {
            _normalize_header(value): position
            for position, value in enumerate(data_header)
            if _normalize_header(value)
        }
        rows: list[dict[str, Any]] = []
        raw_count = 0
        for row_number, row in enumerate(data_sheet.iter_rows(min_row=2, values_only=True), start=2):
            raw_count += 1
            mapped = self._map_row(batch_id=batch_id, source_row_num=row_number, index=data_index, row=row, project_map=project_map)
            if mapped is None:
                continue
            rows.append(mapped)
        self._insert_rows(rows)
        return raw_count, len(rows)

    def _raw(self, index: dict[str, int], row: tuple[object, ...], column_name: str) -> object | None:
        position = index.get(_normalize_header(column_name))
        if position is None or position >= len(row):
            return None
        value = row[position]
        if value is None:
            return None
        if isinstance(value, str) and not value.strip():
            return None
        return value

    def _normalized_age(self, value: object | None) -> int | None:
        age = _to_int(value)
        if age is None or age <= 0:
            return None
        return age

    def _map_row(
        self,
        *,
        batch_id: int,
        source_row_num: int,
        index: dict[str, int],
        row: tuple[object, ...],
        project_map: dict[str, dict[str, str]],
    ) -> dict[str, Any] | None:
        cycle_id = self._raw(index, row, "送检单编号")
        sample_id = self._raw(index, row, "样本编号")
        hospital_name = self._raw(index, row, "送检单位名称")
        review_time = self._raw(index, row, "报告审核时间")
        created_at = _parse_datetime(review_time)
        if created_at is None or not cycle_id or not sample_id or not hospital_name:
            return None

        cycle_id_text = str(cycle_id).strip()
        project_values = project_map.get(cycle_id_text, {})
        sample_type = _normalize_sample_type(self._raw(index, row, "样本类型"))
        values: dict[str, Any] = {
            "import_batch_id": batch_id,
            "source_file_name": self.source_file.name,
            "source_sheet_name": "2025年-数据",
            "source_row_num": source_row_num,
            "imported_at": datetime.utcnow().isoformat(sep=" "),
            "month_bucket": _normalize_month_bucket_value(self._raw(index, row, "月份")) or created_at.strftime("%Y-%m"),
            "report_review_time": _to_iso_datetime(review_time) or created_at.isoformat(sep=" "),
            "hospital_name": str(hospital_name).strip(),
            "product_code": _normalize_text(self._raw(index, row, "产品英文简写")) or "PGT-SR",
            "cycle_id": cycle_id_text,
            "sample_id": str(sample_id).strip(),
            "sample_name": _normalize_text(self._raw(index, row, "送检单样本名称")),
            "sample_type": sample_type,
            "patient_age": self._normalized_age(self._raw(index, row, "受检人年龄")),
            "spouse_age": self._normalized_age(self._raw(index, row, "配偶年龄")),
            "patient_karyotype": _normalize_text(self._raw(index, row, "受检人核型")),
            "spouse_karyotype": _normalize_text(self._raw(index, row, "配偶核型")),
            "data_qc_conclusion": _normalize_qc_conclusion(self._raw(index, row, "data_QC_conclusion")),
            "cnv_result": _normalize_text(self._raw(index, row, "CNV检测结果")),
            "result_label": _normalize_result_label(self._raw(index, row, "结果解释")),
            "cnv_hint": _normalize_text(self._raw(index, row, "提示CNV")),
            "result_detail": _normalize_text(self._raw(index, row, "结果说明")),
            "sr_clinical_type": project_values.get("sr_clinical_type", _normalize_text(self._raw(index, row, "临床指征"))),
            "next_step_screening": project_values.get("next_step_screening", _normalize_text(self._raw(index, row, "是否进入下一步易位筛查"))),
        }
        if "对照" in sample_type or "极体" in sample_type:
            return None
        return values

    def _insert_rows(self, rows: list[dict[str, Any]]) -> None:
        if not rows:
            return
        insert_columns = [name for name, _ in PGTSR_SCHEMA if name != "id"]
        placeholders = ", ".join("?" for _ in insert_columns)
        sql = f"INSERT INTO {PGTSR_TABLE_NAME} ({', '.join(insert_columns)}) VALUES ({placeholders})"
        values = [tuple(row.get(column) for column in insert_columns) for row in rows]
        with _connect(self.database_url) as conn:
            conn.executemany(sql, values)
            conn.commit()


class PGTSRSQLiteRepository:
    def __init__(self, database_url: str) -> None:
        self.database_url = database_url
        self._assert_ready()

    @classmethod
    def from_settings(cls) -> "PGTSRSQLiteRepository":
        return cls(settings.snapshot_db_url)

    def _assert_ready(self) -> None:
        db_path = _sqlite_path_from_url(self.database_url)
        if not db_path.exists():
            raise RuntimeError(
                f"SQLite snapshot database not found: {db_path}. "
                "Run the snapshot builder for PGT-SR first."
            )
        with _connect(self.database_url) as conn:
            exists = conn.execute(
                "SELECT name FROM sqlite_master WHERE type='table' AND name = ?",
                (PGTSR_TABLE_NAME,),
            ).fetchone()
            if exists is None:
                raise RuntimeError("SQLite snapshot database is missing pgtsr_snapshot_raw.")
            count = conn.execute(f"SELECT COUNT(*) FROM {PGTSR_TABLE_NAME}").fetchone()[0]
            if count == 0:
                raise RuntimeError("SQLite snapshot database contains no PGTSR snapshot rows.")

    @property
    def records(self) -> list[PGTSRRecord]:
        return self.filter_records()

    @property
    def eligible_records(self) -> list[PGTSRRecord]:
        return self.filter_records()

    @property
    def stat_month_range(self) -> tuple[str, str]:
        with _connect(self.database_url) as conn:
            row = conn.execute(
                f"""
                SELECT MIN(COALESCE(NULLIF(month_bucket, ''), substr(report_review_time, 1, 7))),
                       MAX(COALESCE(NULLIF(month_bucket, ''), substr(report_review_time, 1, 7)))
                FROM {PGTSR_TABLE_NAME}
                WHERE sample_type NOT LIKE '%对照%' AND sample_type NOT LIKE '%极体%'
                """
            ).fetchone()
        if row and row[0] and row[1]:
            return str(row[0]), str(row[1])
        raise RuntimeError("SQLite snapshot database has no eligible PGTSR month range.")

    @property
    def hospitals(self) -> list[dict[str, str | int]]:
        with _connect(self.database_url) as conn:
            rows = conn.execute(
                f"""
                SELECT hospital_name, COUNT(*) AS sample_count
                FROM {PGTSR_TABLE_NAME}
                WHERE sample_type NOT LIKE '%对照%' AND sample_type NOT LIKE '%极体%'
                GROUP BY hospital_name
                ORDER BY sample_count DESC, hospital_name ASC
                """
            ).fetchall()
        return [
            {"hospital_id": str(row["hospital_name"]), "hospital_name": str(row["hospital_name"]), "sample_count": int(row["sample_count"])}
            for row in rows
        ]

    def filter_records(
        self,
        *,
        hospital_id: str | None = None,
        year: int | None = None,
        quarter: int | None = None,
        month: int | None = None,
        day: int | None = None,
        end_year: int | None = None,
        end_month: int | None = None,
        start_day: date | None = None,
        end_day: date | None = None,
        patient_age_range: str | None = None,
        spouse_age_range: str | None = None,
        sr_clinical_type: str | None = None,
    ) -> list[PGTSRRecord]:
        clauses = ["sample_type NOT LIKE '%对照%'", "sample_type NOT LIKE '%极体%'"]
        params: list[Any] = []
        if hospital_id:
            clauses.append("hospital_name = ?")
            params.append(hospital_id)
        if sr_clinical_type:
            clauses.append("sr_clinical_type = ?")
            params.append(sr_clinical_type)
        if start_day and end_day:
            clauses.append("date(report_review_time) >= date(?)")
            clauses.append("date(report_review_time) <= date(?)")
            params.extend([start_day.isoformat(), end_day.isoformat()])
        elif year is not None:
            if end_year is not None and end_month is not None and month is not None:
                clauses.append(
                    "((CAST(substr(COALESCE(NULLIF(month_bucket, ''), substr(report_review_time, 1, 7)), 1, 4) AS INTEGER) * 12) + "
                    "CAST(substr(COALESCE(NULLIF(month_bucket, ''), substr(report_review_time, 1, 7)), 6, 2) AS INTEGER)) BETWEEN ? AND ?"
                )
                params.extend([year * 12 + month, end_year * 12 + end_month])
            else:
                clauses.append("CAST(substr(COALESCE(NULLIF(month_bucket, ''), substr(report_review_time, 1, 7)), 1, 4) AS INTEGER) = ?")
                params.append(year)
                if quarter is not None:
                    start_month = (quarter - 1) * 3 + 1
                    clauses.append("CAST(substr(COALESCE(NULLIF(month_bucket, ''), substr(report_review_time, 1, 7)), 6, 2) AS INTEGER) BETWEEN ? AND ?")
                    params.extend([start_month, start_month + 2])
                if month is not None:
                    clauses.append("CAST(substr(COALESCE(NULLIF(month_bucket, ''), substr(report_review_time, 1, 7)), 6, 2) AS INTEGER) = ?")
                    params.append(month)
                if day is not None:
                    clauses.append("CAST(substr(date(report_review_time), 9, 2) AS INTEGER) = ?")
                    params.append(day)
        sql = f"SELECT * FROM {PGTSR_TABLE_NAME} WHERE {' AND '.join(clauses)}"
        with _connect(self.database_url) as conn:
            rows = conn.execute(sql, params).fetchall()
        records = [self._row_to_record(row) for row in rows]
        if patient_age_range:
            records = [record for record in records if _match_numeric_age(record.patient_age, patient_age_range)]
        if spouse_age_range:
            records = [record for record in records if _match_numeric_age(record.spouse_age, spouse_age_range)]
        return records

    def _row_to_record(self, row: sqlite3.Row) -> PGTSRRecord:
        created_at = _parse_datetime(row["report_review_time"])
        if created_at is None:
            raise RuntimeError("Invalid report_review_time in SQLite snapshot row.")
        return PGTSRRecord(
            created_at=created_at,
            month_bucket=str(row["month_bucket"] or created_at.strftime("%Y-%m")),
            cycle_id=str(row["cycle_id"] or "").strip(),
            sample_id=str(row["sample_id"] or "").strip(),
            sample_name=str(row["sample_name"] or "").strip(),
            sample_type=_normalize_sample_type(row["sample_type"]),
            hospital_name=str(row["hospital_name"] or "").strip(),
            product_name=str(row["product_code"] or "").strip(),
            result_label=_normalize_result_label(row["result_label"]),
            qc_conclusion=_normalize_qc_conclusion(row["data_qc_conclusion"]),
            cnv_result=_normalize_text(row["cnv_result"]),
            cnv_hint=_normalize_text(row["cnv_hint"]),
            result_detail=_normalize_text(row["result_detail"]),
            patient_age=_to_int(row["patient_age"]),
            spouse_age=_to_int(row["spouse_age"]),
            patient_karyotype=_normalize_text(row["patient_karyotype"]),
            spouse_karyotype=_normalize_text(row["spouse_karyotype"]),
            sr_clinical_type=_normalize_text(row["sr_clinical_type"]),
            next_step_screening=_normalize_text(row["next_step_screening"]),
        )


def _match_numeric_age(age: int | None, age_range: str) -> bool:
    if age_range == "missing":
        return age is None
    if age is None:
        return False
    if age_range.startswith("between:"):
        lower_text, upper_text = age_range.split(":", 1)[1].split(",", 1)
        return int(lower_text) <= age <= int(upper_text)
    if age_range.startswith("gte:"):
        return age >= int(age_range.split(":", 1)[1])
    if age_range.startswith("gt:"):
        return age > int(age_range.split(":", 1)[1])
    if age_range.startswith("lte:"):
        return age <= int(age_range.split(":", 1)[1])
    if age_range.startswith("lt:"):
        return age < int(age_range.split(":", 1)[1])
    return False


def inspect_pgtsr_snapshot_db(database_url: str) -> dict[str, Any]:
    repo = PGTSRSQLiteRepository(database_url)
    with _connect(database_url) as conn:
        row_count = conn.execute(f"SELECT COUNT(*) FROM {PGTSR_TABLE_NAME}").fetchone()[0]
        year_rows = conn.execute(
            f"""
            SELECT substr(COALESCE(NULLIF(month_bucket, ''), substr(report_review_time, 1, 7)), 1, 4) AS snapshot_year,
                   COUNT(*) AS count
            FROM {PGTSR_TABLE_NAME}
            GROUP BY snapshot_year
            ORDER BY snapshot_year
            """
        ).fetchall()
    start, end = repo.stat_month_range
    return {
        "table_name": PGTSR_TABLE_NAME,
        "row_count": int(row_count),
        "years": {int(row["snapshot_year"]): int(row["count"]) for row in year_rows if row["snapshot_year"]},
        "hospital_count": len(repo.hospitals),
        "snapshot_start": start,
        "snapshot_end": end,
    }
