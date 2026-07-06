from __future__ import annotations

import sqlite3
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from yk_review_agent.core.config import settings
from yk_review_agent.services.pgta_detail_dataset import (
    DetailRecord,
    _bucket_month,
    _bucket_quarter,
    _match_age_range,
    _normalize_age_label,
    _normalize_aneuploidy_label,
    _normalize_binary_flag,
    _normalize_qc_conclusion,
    _normalize_result_label,
    _normalize_sample_type,
    _normalize_text,
    _parse_datetime,
    _to_float,
)


PGTA_TABLE_NAME = "pgta_snapshot_raw"
IMPORT_BATCH_TABLE = "snapshot_import_batch"


@dataclass(frozen=True)
class PGTASourceConfig:
    snapshot_year: int
    snapshot_half: str
    file_path: str
    sheet_name: str
    column_map: dict[str, str]


PGTA_SCHEMA: list[tuple[str, str]] = [
    ("id", "INTEGER PRIMARY KEY AUTOINCREMENT"),
    ("snapshot_year", "INTEGER NOT NULL"),
    ("snapshot_half", "TEXT NOT NULL"),
    ("import_batch_id", "INTEGER NOT NULL"),
    ("source_file_name", "TEXT NOT NULL"),
    ("source_sheet_name", "TEXT NOT NULL"),
    ("source_row_num", "INTEGER NOT NULL"),
    ("imported_at", "TEXT NOT NULL"),
    ("month_bucket", "TEXT"),
    ("report_review_time", "TEXT"),
    ("submission_time", "TEXT"),
    ("analysis_time", "TEXT"),
    ("project_type", "TEXT"),
    ("hospital_name", "TEXT"),
    ("hospital_code", "TEXT"),
    ("region_name", "TEXT"),
    ("company_entity", "TEXT"),
    ("product_code", "TEXT"),
    ("order_id", "TEXT"),
    ("order_code", "TEXT"),
    ("sample_barcode", "TEXT"),
    ("sample_id", "TEXT"),
    ("sample_name", "TEXT"),
    ("sample_type", "TEXT"),
    ("charged_sample_count", "INTEGER"),
    ("sample_total_count", "INTEGER"),
    ("received_sample_id", "TEXT"),
    ("case_id", "TEXT"),
    ("doctor_name", "TEXT"),
    ("sales_name", "TEXT"),
    ("is_outsourced", "TEXT"),
    ("payment_method", "TEXT"),
    ("order_source", "TEXT"),
    ("business_type", "TEXT"),
    ("patient_name", "TEXT"),
    ("patient_gender", "TEXT"),
    ("patient_age_system", "INTEGER"),
    ("patient_age_raw", "INTEGER"),
    ("patient_age_manual", "TEXT"),
    ("patient_chromosome", "TEXT"),
    ("patient_karyotype", "TEXT"),
    ("spouse_name", "TEXT"),
    ("spouse_gender", "TEXT"),
    ("spouse_age", "INTEGER"),
    ("spouse_chromosome", "TEXT"),
    ("spouse_karyotype", "TEXT"),
    ("male_father_karyotype", "TEXT"),
    ("male_mother_karyotype", "TEXT"),
    ("female_father_karyotype", "TEXT"),
    ("female_mother_karyotype", "TEXT"),
    ("sample_date", "TEXT"),
    ("received_date", "TEXT"),
    ("input_time", "TEXT"),
    ("sample_status", "TEXT"),
    ("doctor_signed", "TEXT"),
    ("patient_signed", "TEXT"),
    ("order_info_complete", "TEXT"),
    ("transport_status", "TEXT"),
    ("temperature", "TEXT"),
    ("other_karyotype_info", "TEXT"),
    ("morphology_grade", "TEXT"),
    ("adaptation", "TEXT"),
    ("indication_raw", "TEXT"),
    ("indication_system", "TEXT"),
    ("indication_manual", "TEXT"),
    ("other_info", "TEXT"),
    ("resolution", "TEXT"),
    ("family_code", "TEXT"),
    ("amplification_method", "TEXT"),
    ("library_method", "TEXT"),
    ("run_no", "TEXT"),
    ("wga_kit_sku", "TEXT"),
    ("wga_kit_batch", "TEXT"),
    ("wga_concentration", "REAL"),
    ("library_concentration", "REAL"),
    ("cnv_kit_name", "TEXT"),
    ("cnv_kit_batch", "TEXT"),
    ("raw_reads", "REAL"),
    ("high_quality_rate", "REAL"),
    ("mapping_rate", "REAL"),
    ("duplication_rate", "REAL"),
    ("unmapping_rate", "REAL"),
    ("genome_coverage_rate", "REAL"),
    ("valid_reads", "REAL"),
    ("valid_reads_gc_content", "REAL"),
    ("valid_reads_rate", "REAL"),
    ("bin_cv_1000k", "REAL"),
    ("seg_dd", "REAL"),
    ("mt_cn", "REAL"),
    ("data_qc_conclusion", "TEXT"),
    ("data_qc_information", "TEXT"),
    ("sex_karyotype", "TEXT"),
    ("cnv_result", "TEXT"),
    ("result_label", "TEXT"),
    ("cnv_hint", "TEXT"),
    ("chromosome_location", "TEXT"),
    ("result_detail", "TEXT"),
    ("aneuploidy_result_raw", "TEXT"),
    ("aneuploidy_manual", "TEXT"),
    ("incidental_raw", "TEXT"),
    ("incidental_manual", "TEXT"),
    ("result_note_1", "TEXT"),
    ("karyotype_abnormal_flag", "TEXT"),
    ("is_new_process", "TEXT"),
]


def _normalize_header(value: object) -> str:
    return str(value or "").replace("\n", "").strip()


def _to_int(value: object) -> int | None:
    if value in (None, ""):
        return None
    try:
        return int(float(value))
    except (TypeError, ValueError):
        return None


def _to_iso_datetime(value: object) -> str | None:
    dt = _parse_datetime(value)
    return dt.isoformat(sep=" ") if dt else None


def _normalize_month_bucket_value(value: object) -> str | None:
    if value in (None, ""):
        return None
    text = str(value).strip()
    if not text:
        return None
    normalized = text.replace("/", "-").replace("年", "-").replace("月", "").strip()
    if len(normalized) == 4 and normalized.isdigit():
        year = int(normalized[:2])
        month = int(normalized[2:])
        return f"20{year:02d}-{month:02d}"
    if len(normalized) == 7 and normalized[4] == "-" and normalized[:4].isdigit() and normalized[5:].isdigit():
        return f"{int(normalized[:4]):04d}-{int(normalized[5:]):02d}"
    if len(normalized) == 6 and normalized.isdigit():
        return f"{int(normalized[:4]):04d}-{int(normalized[4:]):02d}"
    return text


def _sqlite_path_from_url(database_url: str) -> Path:
    prefix = "sqlite+pysqlite:///"
    if not database_url.startswith(prefix):
        raise RuntimeError(f"Unsupported SNAPSHOT_DB_URL: {database_url}")
    return Path(database_url[len(prefix) :]).resolve()


def _connect(database_url: str) -> sqlite3.Connection:
    path = _sqlite_path_from_url(database_url)
    path.parent.mkdir(parents=True, exist_ok=True)
    conn = sqlite3.connect(path)
    conn.row_factory = sqlite3.Row
    return conn


def _default_pgta_source_configs() -> list[PGTASourceConfig]:
    return [
        PGTASourceConfig(
            snapshot_year=2023,
            snapshot_half="h2",
            file_path=settings.pgta_snapshot_2023_file,
            sheet_name="7月-12月",
            column_map={
                "报告审核时间": "report_review_time",
                "项目类型": "project_type",
                "公司主体": "company_entity",
                "送检单位名称": "hospital_name",
                "收样区域": "region_name",
                "产品英文简写": "product_code",
                "送检单编号": "order_id",
                "送检码": "order_code",
                "收费样本数量": "charged_sample_count",
                "样本总数量": "sample_total_count",
                "送检医生": "doctor_name",
                "销售员": "sales_name",
                "是否外送": "is_outsourced",
                "回款方式": "payment_method",
                "订单来源": "order_source",
                "业务类型": "business_type",
                "受检人姓名": "patient_name",
                "受检人性别": "patient_gender",
                "受检人年龄": "patient_age_raw",
                "受检人染色体": "patient_chromosome",
                "受检人核型": "patient_karyotype",
                "配偶姓名": "spouse_name",
                "配偶性别": "spouse_gender",
                "配偶年龄": "spouse_age",
                "配偶染色体": "spouse_chromosome",
                "配偶核型": "spouse_karyotype",
                "采样日期": "sample_date",
                "样本状态": "sample_status",
                "医生是否签字确认": "doctor_signed",
                "患者是否签字": "patient_signed",
                "送检单信息是否完整": "order_info_complete",
                "运输状态": "transport_status",
                "温度": "temperature",
                "其他染色体/核型信息": "other_karyotype_info",
                "男方父亲核型": "male_father_karyotype",
                "男方母亲核型": "male_mother_karyotype",
                "女方父亲核型": "female_father_karyotype",
                "女方母亲核型": "female_mother_karyotype",
                "收样日期": "received_date",
                "录入时间": "input_time",
                "样本编号": "sample_id",
                "送检单样本名称": "sample_name",
                "样本类型": "sample_type",
                "收样编号": "received_sample_id",
                "形态学评级": "morphology_grade",
                "病例ID": "case_id",
                "适应症": "adaptation",
                "临床指征": "indication_raw",
                "其他信息": "other_info",
                "分辨率": "resolution",
                "扩增方法": "amplification_method",
                "建库方法": "library_method",
                "RUN号": "run_no",
                "WGA扩增试剂盒货号": "wga_kit_sku",
                "WGA扩增试剂盒批号": "wga_kit_batch",
                "检测浓度（WGA扩增浓度）": "wga_concentration",
                "建库文库浓度（酶切打断文库浓度）": "library_concentration",
                "CNV试剂盒名称": "cnv_kit_name",
                "CNV试剂盒批次": "cnv_kit_batch",
                "raw_reads": "raw_reads",
                "high_quality_rate(%)": "high_quality_rate",
                "mapping_rate(%)": "mapping_rate",
                "duplication_rate(%)": "duplication_rate",
                "coverage_of_genome(%)": "genome_coverage_rate",
                "valid_reads": "valid_reads",
                "valid_reads_GC_content(%)": "valid_reads_gc_content",
                "valid_reads_rate(%)": "valid_reads_rate",
                "CV(1000K_bin_size)": "bin_cv_1000k",
                "data_QC_conclusion": "data_qc_conclusion",
                "性染色体核型": "sex_karyotype",
                "CNV检测结果": "cnv_result",
                "结果解释": "result_label",
                "提示CNV": "cnv_hint",
                "染色体位置": "chromosome_location",
                "结果说明": "result_detail",
                "异倍体分析（对应亲本污染分析字段）": "aneuploidy_result_raw",
                "结果备注1": "result_note_1",
            },
        ),
        PGTASourceConfig(
            snapshot_year=2024,
            snapshot_half="full",
            file_path=settings.pgta_snapshot_2024_file,
            sheet_name="源数据",
            column_map={
                "月份": "month_bucket",
                "报告审核时间": "report_review_time",
                "项目类型": "project_type",
                "公司主体": "company_entity",
                "送检单位名称": "hospital_name",
                "收样区域": "region_name",
                "产品英文简写": "product_code",
                "送检单编号": "order_id",
                "送检码": "order_code",
                "收费样本数量": "charged_sample_count",
                "样本总数量": "sample_total_count",
                "送检医生": "doctor_name",
                "销售员": "sales_name",
                "是否外送": "is_outsourced",
                "回款方式": "payment_method",
                "订单来源": "order_source",
                "业务类型": "business_type",
                "受检人姓名": "patient_name",
                "受检人性别": "patient_gender",
                "受检人年龄": "patient_age_raw",
                "受检人染色体": "patient_chromosome",
                "受检人核型": "patient_karyotype",
                "配偶姓名": "spouse_name",
                "配偶性别": "spouse_gender",
                "配偶年龄": "spouse_age",
                "配偶染色体": "spouse_chromosome",
                "配偶核型": "spouse_karyotype",
                "采样日期": "sample_date",
                "样本状态": "sample_status",
                "医生是否签字确认": "doctor_signed",
                "患者是否签字": "patient_signed",
                "送检单信息是否完整": "order_info_complete",
                "运输状态": "transport_status",
                "温度": "temperature",
                "其他染色体/核型信息": "other_karyotype_info",
                "男方父亲核型": "male_father_karyotype",
                "男方母亲核型": "male_mother_karyotype",
                "女方父亲核型": "female_father_karyotype",
                "女方母亲核型": "female_mother_karyotype",
                "收样日期": "received_date",
                "录入时间": "input_time",
                "样本编号": "sample_id",
                "送检单样本名称": "sample_name",
                "样本类型": "sample_type",
                "收样编号": "received_sample_id",
                "形态学评级": "morphology_grade",
                "病例ID": "case_id",
                "适应症": "adaptation",
                "临床指征": "indication_raw",
                "其他信息": "other_info",
                "分辨率": "resolution",
                "家系代码": "family_code",
                "扩增方法": "amplification_method",
                "建库方法": "library_method",
                "RUN号": "run_no",
                "WGA扩增试剂盒货号": "wga_kit_sku",
                "WGA扩增试剂盒批号": "wga_kit_batch",
                "检测浓度（WGA扩增浓度）": "wga_concentration",
                "建库文库浓度（酶切打断文库浓度）": "library_concentration",
                "CNV试剂盒名称": "cnv_kit_name",
                "CNV试剂盒批次": "cnv_kit_batch",
                "raw_reads": "raw_reads",
                "high_quality_rate(%)": "high_quality_rate",
                "mapping_rate(%)": "mapping_rate",
                "duplication_rate(%)": "duplication_rate",
                "coverage_of_genome(%)": "genome_coverage_rate",
                "valid_reads": "valid_reads",
                "valid_reads_GC_content(%)": "valid_reads_gc_content",
                "valid_reads_rate(%)": "valid_reads_rate",
                "CV(1000K_bin_size)": "bin_cv_1000k",
                "data_QC_conclusion": "data_qc_conclusion",
                "性染色体核型": "sex_karyotype",
                "CNV检测结果": "cnv_result",
                "结果解释": "result_label",
                "提示CNV": "cnv_hint",
                "染色体位置": "chromosome_location",
                "结果说明": "result_detail",
                "异倍体分析（对应亲本污染分析字段）": "aneuploidy_result_raw",
                "结果备注1": "result_note_1",
                "异倍体分析": "aneuploidy_manual",
                "是否有意外发现": "incidental_raw",
            },
        ),
        PGTASourceConfig(
            snapshot_year=2025,
            snapshot_half="full",
            file_path=settings.pgta_snapshot_2025_file,
            sheet_name="2025年-数据",
            column_map={},
        ),
        PGTASourceConfig(
            snapshot_year=2026,
            snapshot_half="full",
            file_path=settings.pgta_snapshot_2026_file,
            sheet_name="2026年-数据",
            column_map={},
        ),
    ]


PGTA_BASE_COLUMN_MAP: dict[str, str] = {
    "月份": "month_bucket",
    "报告审核时间": "report_review_time",
    "项目类型": "project_type",
    "公司主体": "company_entity",
    "送检单位名称": "hospital_name",
    "收样区域": "region_name",
    "产品英文简写": "product_code",
    "送检单编号": "order_id",
    "送检码": "order_code",
    "收费样本数量": "charged_sample_count",
    "样本总数量": "sample_total_count",
    "送检医生": "doctor_name",
    "销售员": "sales_name",
    "是否外送": "is_outsourced",
    "回款方式": "payment_method",
    "订单来源": "order_source",
    "业务类型": "business_type",
    "受检人姓名": "patient_name",
    "受检人性别": "patient_gender",
    "受检人年龄（系统）": "patient_age_system",
    "受检人年龄": "patient_age_raw",
    "受检人染色体": "patient_chromosome",
    "受检人核型": "patient_karyotype",
    "配偶姓名": "spouse_name",
    "配偶性别": "spouse_gender",
    "配偶年龄": "spouse_age",
    "配偶染色体": "spouse_chromosome",
    "配偶核型": "spouse_karyotype",
    "采样日期": "sample_date",
    "样本状态": "sample_status",
    "医生是否签字确认": "doctor_signed",
    "患者是否签字": "patient_signed",
    "送检单信息是否完整": "order_info_complete",
    "运输状态": "transport_status",
    "温度": "temperature",
    "其他染色体/核型信息": "other_karyotype_info",
    "男方父亲核型": "male_father_karyotype",
    "男方母亲核型": "male_mother_karyotype",
    "女方父亲核型": "female_father_karyotype",
    "女方母亲核型": "female_mother_karyotype",
    "收样日期": "received_date",
    "录入时间": "input_time",
    "样本编号": "sample_id",
    "送检单样本名称": "sample_name",
    "样本类型": "sample_type",
    "收样编号": "received_sample_id",
    "形态学评级": "morphology_grade",
    "病例ID": "case_id",
    "适应症": "adaptation",
    "临床指征（系统）": "indication_system",
    "临床指征": "indication_raw",
    "其他信息": "other_info",
    "分辨率": "resolution",
    "家系代码": "family_code",
    "扩增方法": "amplification_method",
    "建库方法": "library_method",
    "RUN号": "run_no",
    "WGA扩增试剂盒货号": "wga_kit_sku",
    "WGA扩增试剂盒批号": "wga_kit_batch",
    "检测浓度（WGA扩增浓度）": "wga_concentration",
    "建库文库浓度（酶切打断文库浓度）": "library_concentration",
    "CNV试剂盒名称": "cnv_kit_name",
    "CNV试剂盒批次": "cnv_kit_batch",
    "raw_reads": "raw_reads",
    "high_quality_rate(%)": "high_quality_rate",
    "mapping_rate(%)": "mapping_rate",
    "duplication_rate(%)": "duplication_rate",
    "coverage_of_genome(%)": "genome_coverage_rate",
    "valid_reads": "valid_reads",
    "valid_reads_GC_content(%)": "valid_reads_gc_content",
    "valid_reads_rate(%)": "valid_reads_rate",
    "CV(1000K_bin_size)": "bin_cv_1000k",
    "data_QC_conclusion": "data_qc_conclusion",
    "性染色体核型": "sex_karyotype",
    "CNV检测结果": "cnv_result",
    "结果解释": "result_label",
    "提示CNV": "cnv_hint",
    "染色体位置": "chromosome_location",
    "结果说明": "result_detail",
    "异倍体分析（对应亲本污染分析字段）": "aneuploidy_result_raw",
    "结果备注1": "result_note_1",
    "异倍体分析": "aneuploidy_manual",
    "是否有意外发现": "incidental_raw",
    "受检人年龄（人工处理）": "patient_age_manual",
    "临床指征（人工处理）": "indication_manual",
    "意外发现（人工处理）": "incidental_manual",
    "异倍体结果（人工处理）": "aneuploidy_manual",
}


def _build_column_map(config: PGTASourceConfig) -> dict[str, str]:
    merged = dict(PGTA_BASE_COLUMN_MAP)
    merged.update(config.column_map)
    return merged


def create_snapshot_schema(database_url: str) -> None:
    columns = ", ".join(f"{name} {definition}" for name, definition in PGTA_SCHEMA)
    with _connect(database_url) as conn:
        conn.execute(
            f"""
            CREATE TABLE IF NOT EXISTS {IMPORT_BATCH_TABLE} (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                product_code TEXT NOT NULL,
                table_name TEXT NOT NULL,
                source_file_name TEXT NOT NULL,
                source_sheet_name TEXT NOT NULL,
                source_file_size INTEGER NOT NULL,
                source_file_mtime TEXT NOT NULL,
                import_started_at TEXT NOT NULL,
                import_finished_at TEXT,
                row_count_raw INTEGER NOT NULL DEFAULT 0,
                row_count_loaded INTEGER NOT NULL DEFAULT 0,
                status TEXT NOT NULL,
                notes TEXT NOT NULL DEFAULT ''
            )
            """
        )
        conn.execute(f"CREATE TABLE IF NOT EXISTS {PGTA_TABLE_NAME} ({columns})")
        conn.execute(f"CREATE INDEX IF NOT EXISTS idx_pgta_hospital ON {PGTA_TABLE_NAME}(hospital_name)")
        conn.execute(f"CREATE INDEX IF NOT EXISTS idx_pgta_review_time ON {PGTA_TABLE_NAME}(report_review_time)")
        conn.execute(f"CREATE INDEX IF NOT EXISTS idx_pgta_month_bucket ON {PGTA_TABLE_NAME}(month_bucket)")
        conn.execute(f"CREATE INDEX IF NOT EXISTS idx_pgta_order_id ON {PGTA_TABLE_NAME}(order_id)")
        conn.execute(f"CREATE INDEX IF NOT EXISTS idx_pgta_sample_id ON {PGTA_TABLE_NAME}(sample_id)")
        conn.execute(
            f"CREATE INDEX IF NOT EXISTS idx_pgta_hospital_month ON {PGTA_TABLE_NAME}(hospital_name, month_bucket)"
        )


class PGTAExcelImporter:
    def __init__(self, database_url: str, *, source_configs: list[PGTASourceConfig] | None = None) -> None:
        self.database_url = database_url
        self.source_configs = source_configs or _default_pgta_source_configs()

    def rebuild(self) -> None:
        create_snapshot_schema(self.database_url)
        with _connect(self.database_url) as conn:
            conn.execute(f"DELETE FROM {PGTA_TABLE_NAME}")
            conn.execute(f"DELETE FROM {IMPORT_BATCH_TABLE} WHERE product_code = 'PGT-A'")
            conn.commit()
        for config in self.source_configs:
            self._import_source(config)

    def _import_source(self, config: PGTASourceConfig) -> None:
        source_path = Path(config.file_path).resolve()
        if not source_path.exists():
            raise RuntimeError(f"PGTA source file not found: {source_path}")
        started_at = datetime.utcnow().isoformat(sep=" ")
        with _connect(self.database_url) as conn:
            cursor = conn.execute(
                f"""
                INSERT INTO {IMPORT_BATCH_TABLE} (
                    product_code, table_name, source_file_name, source_sheet_name,
                    source_file_size, source_file_mtime, import_started_at, status
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    "PGT-A",
                    PGTA_TABLE_NAME,
                    source_path.name,
                    config.sheet_name,
                    source_path.stat().st_size,
                    datetime.fromtimestamp(source_path.stat().st_mtime).isoformat(sep=" "),
                    started_at,
                    "running",
                ),
            )
            batch_id = int(cursor.lastrowid)
            conn.commit()
        try:
            row_count_raw, row_count_loaded = self._import_rows(config, batch_id)
            status = "success"
            notes = ""
        except Exception as exc:
            row_count_raw = 0
            row_count_loaded = 0
            status = "failed"
            notes = str(exc)
            raise
        finally:
            with _connect(self.database_url) as conn:
                conn.execute(
                    f"""
                    UPDATE {IMPORT_BATCH_TABLE}
                    SET import_finished_at = ?, row_count_raw = ?, row_count_loaded = ?, status = ?, notes = ?
                    WHERE id = ?
                    """,
                    (
                        datetime.utcnow().isoformat(sep=" "),
                        row_count_raw,
                        row_count_loaded,
                        status,
                        notes,
                        batch_id,
                    ),
                )
                conn.commit()

    def _import_rows(self, config: PGTASourceConfig, batch_id: int) -> tuple[int, int]:
        workbook = load_workbook(config.file_path, read_only=True, data_only=True)
        if config.sheet_name not in workbook.sheetnames:
            raise RuntimeError(f"Sheet not found: {config.sheet_name} in {config.file_path}")
        sheet = workbook[config.sheet_name]
        header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
        index = {
            _normalize_header(value): position
            for position, value in enumerate(header_row)
            if _normalize_header(value)
        }
        column_map = _build_column_map(config)
        raw_count = 0
        loaded_rows: list[dict[str, Any]] = []
        for row_number, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            raw_count += 1
            mapped = self._map_row(config, batch_id, source_row_num=row_number, index=index, row=row)
            if mapped is None:
                continue
            loaded_rows.append(mapped)
        self._insert_rows(loaded_rows)
        return raw_count, len(loaded_rows)

    def _map_row(
        self,
        config: PGTASourceConfig,
        batch_id: int,
        *,
        source_row_num: int,
        index: dict[str, int],
        row: tuple[object, ...],
    ) -> dict[str, Any] | None:
        column_map = _build_column_map(config)

        def raw(column_name: str) -> object | None:
            position = index.get(_normalize_header(column_name))
            if position is None or position >= len(row):
                return None
            value = row[position]
            if value is None:
                return None
            if isinstance(value, str) and not value.strip():
                return None
            return value

        values: dict[str, Any] = {
            "snapshot_year": config.snapshot_year,
            "snapshot_half": config.snapshot_half,
            "import_batch_id": batch_id,
            "source_file_name": Path(config.file_path).name,
            "source_sheet_name": config.sheet_name,
            "source_row_num": source_row_num,
            "imported_at": datetime.utcnow().isoformat(sep=" "),
        }
        for source_name, target_name in column_map.items():
            values[target_name] = raw(source_name)

        report_review_time = values.get("report_review_time") or values.get("analysis_time")
        submission_time = values.get("submission_time")
        created_at = _parse_datetime(report_review_time) or _parse_datetime(submission_time)
        sample_id = values.get("sample_id") or values.get("sample_barcode")
        hospital_name = values.get("hospital_name")
        if created_at is None or not hospital_name or not sample_id:
            return None

        values["report_review_time"] = values.get("report_review_time") or created_at.isoformat(sep=" ")
        values["month_bucket"] = _normalize_month_bucket_value(values.get("month_bucket")) or created_at.strftime("%Y-%m")
        values["sample_id"] = str(sample_id).strip()
        values["hospital_name"] = str(hospital_name).strip()
        values["product_code"] = values.get("product_code") or "PGT-A"
        values["sample_type"] = _normalize_sample_type(values.get("sample_type"))
        values["result_label"] = _normalize_result_label(values.get("result_label"))
        values["data_qc_conclusion"] = _normalize_qc_conclusion(values.get("data_qc_conclusion"))
        values["cnv_result"] = _normalize_text(values.get("cnv_result"))
        values["cnv_hint"] = _normalize_text(values.get("cnv_hint"))
        values["chromosome_location"] = _normalize_text(values.get("chromosome_location"))
        values["result_detail"] = _normalize_text(values.get("result_detail"))
        values["incidental_raw"] = _normalize_binary_flag(values.get("incidental_raw"))
        values["incidental_manual"] = _normalize_binary_flag(values.get("incidental_manual"))
        values["aneuploidy_manual"] = _normalize_aneuploidy_label(values.get("aneuploidy_manual"))
        values["aneuploidy_result_raw"] = _normalize_aneuploidy_label(values.get("aneuploidy_result_raw"))

        for key in (
            "patient_age_system",
            "patient_age_raw",
            "spouse_age",
            "charged_sample_count",
            "sample_total_count",
        ):
            values[key] = _to_int(values.get(key))

        for key in (
            "wga_concentration",
            "library_concentration",
            "raw_reads",
            "high_quality_rate",
            "mapping_rate",
            "duplication_rate",
            "unmapping_rate",
            "genome_coverage_rate",
            "valid_reads",
            "valid_reads_gc_content",
            "valid_reads_rate",
            "bin_cv_1000k",
            "seg_dd",
            "mt_cn",
        ):
            values[key] = _to_float(values.get(key))

        for key in ("report_review_time", "submission_time", "analysis_time", "sample_date", "received_date", "input_time"):
            values[key] = _to_iso_datetime(values.get(key))
        return values

    def _insert_rows(self, rows: list[dict[str, Any]]) -> None:
        if not rows:
            return
        insert_columns = [name for name, _ in PGTA_SCHEMA if name != "id"]
        placeholders = ", ".join("?" for _ in insert_columns)
        sql = f"INSERT INTO {PGTA_TABLE_NAME} ({', '.join(insert_columns)}) VALUES ({placeholders})"
        values = [tuple(row.get(column) for column in insert_columns) for row in rows]
        with _connect(self.database_url) as conn:
            conn.executemany(sql, values)
            conn.commit()


class PGTASQLiteRepository:
    def __init__(self, database_url: str) -> None:
        self.database_url = database_url
        self._assert_ready()

    @classmethod
    def from_settings(cls) -> "PGTASQLiteRepository":
        return cls(settings.snapshot_db_url)

    def _assert_ready(self) -> None:
        db_path = _sqlite_path_from_url(self.database_url)
        if not db_path.exists():
            raise RuntimeError(
                f"SQLite snapshot database not found: {db_path}. "
                "Run `python -m yk_review_agent.tools.build_snapshot_db --product pgta` first."
            )
        with _connect(self.database_url) as conn:
            exists = conn.execute(
                "SELECT name FROM sqlite_master WHERE type='table' AND name = ?",
                (PGTA_TABLE_NAME,),
            ).fetchone()
            if exists is None:
                raise RuntimeError("SQLite snapshot database is missing pgta_snapshot_raw.")
            count = conn.execute(f"SELECT COUNT(*) FROM {PGTA_TABLE_NAME}").fetchone()[0]
            if count == 0:
                raise RuntimeError("SQLite snapshot database contains no PGTA snapshot rows.")

    @property
    def records(self) -> list[DetailRecord]:
        return self.filter_records()

    @property
    def eligible_records(self) -> list[DetailRecord]:
        return self.filter_records()

    @property
    def stat_month_range(self) -> tuple[str, str]:
        with _connect(self.database_url) as conn:
            row = conn.execute(
                f"""
                SELECT MIN(COALESCE(NULLIF(month_bucket, ''), substr(report_review_time, 1, 7))),
                       MAX(COALESCE(NULLIF(month_bucket, ''), substr(report_review_time, 1, 7)))
                FROM {PGTA_TABLE_NAME}
                WHERE sample_type NOT LIKE '%对照%' AND sample_type NOT LIKE '%极体%'
                """
            ).fetchone()
        if row and row[0] and row[1]:
            return str(row[0]), str(row[1])
        raise RuntimeError("SQLite snapshot database has no eligible PGTA month range.")

    @property
    def hospitals(self) -> list[dict[str, str | int]]:
        with _connect(self.database_url) as conn:
            rows = conn.execute(
                f"""
                SELECT hospital_name, COUNT(*) AS sample_count
                FROM {PGTA_TABLE_NAME}
                WHERE sample_type NOT LIKE '%对照%' AND sample_type NOT LIKE '%极体%'
                GROUP BY hospital_name
                ORDER BY sample_count DESC, hospital_name ASC
                """
            ).fetchall()
        return [
            {"hospital_id": str(row["hospital_name"]), "hospital_name": str(row["hospital_name"]), "sample_count": int(row["sample_count"])}
            for row in rows
        ]

    def filter_records(
        self,
        *,
        hospital_id: str | None = None,
        year: int | None = None,
        quarter: int | None = None,
        month: int | None = None,
        day: int | None = None,
        end_year: int | None = None,
        end_month: int | None = None,
        start_day: date | None = None,
        end_day: date | None = None,
        age_range: str | None = None,
    ) -> list[DetailRecord]:
        clauses = ["sample_type NOT LIKE '%对照%'", "sample_type NOT LIKE '%极体%'"]
        params: list[Any] = []
        if hospital_id:
            clauses.append("hospital_name = ?")
            params.append(hospital_id)
        if start_day and end_day:
            clauses.append("date(report_review_time) >= date(?)")
            clauses.append("date(report_review_time) <= date(?)")
            params.extend([start_day.isoformat(), end_day.isoformat()])
        elif year is not None:
            if end_year is not None and end_month is not None and month is not None:
                clauses.append(
                    "((CAST(substr(COALESCE(NULLIF(month_bucket, ''), substr(report_review_time, 1, 7)), 1, 4) AS INTEGER) * 12) + "
                    "CAST(substr(COALESCE(NULLIF(month_bucket, ''), substr(report_review_time, 1, 7)), 6, 2) AS INTEGER)) BETWEEN ? AND ?"
                )
                params.extend([year * 12 + month, end_year * 12 + end_month])
            else:
                clauses.append("CAST(substr(COALESCE(NULLIF(month_bucket, ''), substr(report_review_time, 1, 7)), 1, 4) AS INTEGER) = ?")
                params.append(year)
                if quarter is not None:
                    start_month = (quarter - 1) * 3 + 1
                    clauses.append("CAST(substr(COALESCE(NULLIF(month_bucket, ''), substr(report_review_time, 1, 7)), 6, 2) AS INTEGER) BETWEEN ? AND ?")
                    params.extend([start_month, start_month + 2])
                if month is not None:
                    clauses.append("CAST(substr(COALESCE(NULLIF(month_bucket, ''), substr(report_review_time, 1, 7)), 6, 2) AS INTEGER) = ?")
                    params.append(month)
                if day is not None:
                    clauses.append("CAST(substr(date(report_review_time), 9, 2) AS INTEGER) = ?")
                    params.append(day)
        sql = f"SELECT * FROM {PGTA_TABLE_NAME} WHERE {' AND '.join(clauses)}"
        with _connect(self.database_url) as conn:
            rows = conn.execute(sql, params).fetchall()
        records = [self._row_to_record(row) for row in rows]
        if age_range:
            records = [record for record in records if _match_age_range(record.age, record.age_label, age_range)]
        if quarter is not None and year is None:
            records = [record for record in records if _bucket_quarter(record.month_bucket, record.created_at) == quarter]
        if month is not None and year is None and end_month is None:
            records = [record for record in records if _bucket_month(record.month_bucket, record.created_at) == month]
        if day is not None and year is None and not (start_day and end_day):
            records = [record for record in records if record.created_at.day == day]
        return records

    def _row_to_record(self, row: sqlite3.Row) -> DetailRecord:
        created_at = _parse_datetime(row["report_review_time"])
        if created_at is None:
            raise RuntimeError("Invalid report_review_time in SQLite snapshot row.")
        age = row["patient_age_system"]
        if age is None:
            age = row["patient_age_raw"]
        age_label = _normalize_age_label(row["patient_age_manual"])
        return DetailRecord(
            created_at=created_at,
            month_bucket=str(row["month_bucket"] or created_at.strftime("%Y-%m")),
            cycle_id=str(row["order_id"] or "").strip(),
            task_name=str(row["order_id"] or "").strip(),
            sample_id=str(row["sample_id"] or "").strip(),
            sample_name=str(row["sample_name"] or "").strip(),
            sample_type=_normalize_sample_type(row["sample_type"]),
            hospital_name=str(row["hospital_name"] or "").strip(),
            hospital_code=str(row["hospital_name"] or "").strip(),
            product_name=str(row["product_code"] or "").strip(),
            description_cn=_normalize_result_label(row["result_label"]),
            description_en="",
            qc_conclusion=_normalize_qc_conclusion(row["data_qc_conclusion"]),
            ldpgta_qc_conclusion="",
            age=_to_int(age),
            age_label=age_label,
            mapd=None,
            bincv=_to_float(row["bin_cv_1000k"]),
            cnvpq=None,
            cnv_type=_normalize_text(row["cnv_result"]),
            upd_arms=_normalize_text(row["chromosome_location"]),
            ldpgta_cnv_type=_normalize_text(row["cnv_hint"]),
            result_detail=_normalize_text(row["result_detail"]),
            incidental_label=_normalize_binary_flag(row["incidental_manual"] or row["incidental_raw"]),
            aneuploidy_label=_normalize_aneuploidy_label(row["aneuploidy_manual"] or row["aneuploidy_result_raw"]),
        )


def inspect_pgta_snapshot_db(database_url: str) -> dict[str, Any]:
    repo = PGTASQLiteRepository(database_url)
    with _connect(database_url) as conn:
        row_count = conn.execute(f"SELECT COUNT(*) FROM {PGTA_TABLE_NAME}").fetchone()[0]
        year_rows = conn.execute(
            f"SELECT snapshot_year, COUNT(*) AS count FROM {PGTA_TABLE_NAME} GROUP BY snapshot_year ORDER BY snapshot_year"
        ).fetchall()
    start, end = repo.stat_month_range
    return {
        "table_name": PGTA_TABLE_NAME,
        "row_count": int(row_count),
        "years": {int(row["snapshot_year"]): int(row["count"]) for row in year_rows},
        "hospital_count": len(repo.hospitals),
        "snapshot_start": start,
        "snapshot_end": end,
    }
