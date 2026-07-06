from __future__ import annotations

from collections import Counter
from dataclasses import dataclass
from datetime import date, datetime
from functools import lru_cache
from pathlib import Path
import re

from openpyxl import load_workbook

from yk_review_agent.core.config import settings


@dataclass(frozen=True)
class DetailRecord:
    created_at: datetime
    month_bucket: str
    cycle_id: str
    task_name: str
    sample_id: str
    sample_name: str
    sample_type: str
    hospital_name: str
    hospital_code: str
    product_name: str
    description_cn: str
    description_en: str
    qc_conclusion: str
    ldpgta_qc_conclusion: str
    age: int | None
    age_label: str
    mapd: float | None
    bincv: float | None
    cnvpq: float | None
    cnv_type: str
    upd_arms: str
    ldpgta_cnv_type: str
    result_detail: str
    incidental_label: str
    aneuploidy_label: str

    @property
    def is_embryo_sample(self) -> bool:
        sample_type = self.sample_type or ""
        return "对照" not in sample_type and "极体" not in sample_type

    @property
    def is_qc_fail(self) -> bool:
        return self.description_cn == "质控不合格" or self.qc_conclusion == "FAIL"

    @property
    def has_main_cnv_result(self) -> bool:
        text = (self.cnv_type or "").strip()
        return bool(text and text not in {"/", "N/A"})

    @property
    def is_euploid(self) -> bool:
        return self.description_cn == "未见异常"

    @property
    def is_mosaic(self) -> bool:
        return self.description_cn == "嵌合异常"

    @property
    def is_abnormal(self) -> bool:
        return self.description_cn == "异常"


def _parse_datetime(value: object) -> datetime | None:
    if isinstance(value, datetime):
        return value
    if isinstance(value, date):
        return datetime.combine(value, datetime.min.time())
    if value in (None, ""):
        return None
    text = str(value).strip()
    if not text:
        return None
    for fmt in (
        "%Y-%m-%d %H:%M:%S",
        "%Y-%m-%d",
        "%Y/%m/%d %H:%M:%S",
        "%Y/%m/%d",
        "%Y年%m月%d日",
    ):
        try:
            return datetime.strptime(text, fmt)
        except ValueError:
            continue
    return None


def _parse_age(value: object) -> int | None:
    if value in (None, "", 0, "0"):
        return None
    try:
        return int(float(value))
    except (TypeError, ValueError):
        return None


def _normalize_age_label(value: object) -> str:
    text = str(value or "").strip()
    if not text:
        return "未填写"
    mapping = {
        "<35岁": "＜35岁",
        "35-37岁": "35-38岁",
        "35-38岁": "35-38岁",
        ">38岁": "＞38岁",
        ">=39岁": "＞38岁",
        "＞38岁": "＞38岁",
        "未填写": "未填写",
    }
    return mapping.get(text, text)


def _normalize_text(value: object) -> str:
    return str(value or "").strip()


def _normalize_sample_type(value: object) -> str:
    text = _normalize_text(value)
    return re.sub(r"\s+", " ", text)


def _normalize_result_label(value: object) -> str:
    text = _normalize_text(value)
    mapping = {
        "未见染色体拷贝数异常": "未见异常",
        "—": "/",
        "——": "/",
        "-": "/",
    }
    return mapping.get(text, text)


def _normalize_qc_conclusion(value: object) -> str:
    text = _normalize_text(value).upper()
    if text in {"/", "-", "—", "——", "XX"}:
        return ""
    return text


def _normalize_binary_flag(value: object) -> str:
    text = _normalize_text(value)
    mapping = {
        "是": "有",
        "否": "无",
        "/": "",
        "-": "",
        "—": "",
        "——": "",
    }
    return mapping.get(text, text)


def _normalize_aneuploidy_label(value: object) -> str:
    text = _normalize_text(value)
    if text in {"/", "-", "—", "——"}:
        return ""
    return text


def _to_float(value: object) -> float | None:
    if value in (None, ""):
        return None
    if isinstance(value, str) and value.strip() in {"/", "-"}:
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


class PGTADetailDataset:
    def __init__(self, file_path: str) -> None:
        self.file_path = Path(file_path)
        self.sheet_name = "2025年-数据"
        self._records = self._load_records()

    def _load_records(self) -> list[DetailRecord]:
        workbook = load_workbook(self.file_path, data_only=True, read_only=True)
        sheet_name = self.sheet_name if self.sheet_name in workbook.sheetnames else workbook.sheetnames[0]
        sheet = workbook[sheet_name]
        header = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
        index = {
            str(value).strip(): position
            for position, value in enumerate(header)
            if value is not None and str(value).strip()
        }

        def value(row: tuple[object, ...], *columns: str) -> object:
            for column in columns:
                position = index.get(column)
                if position is None:
                    continue
                current = row[position]
                if current is None:
                    continue
                if isinstance(current, str) and not current.strip():
                    continue
                return current
            return None

        records: list[DetailRecord] = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            created_at = _parse_datetime(value(row, "报告审核时间", "项目日期"))
            sample_id = str(value(row, "样本编号") or "").strip()
            hospital_name = str(value(row, "送检单位名称") or "").strip()
            sample_type = _normalize_sample_type(value(row, "样本类型"))
            month_bucket = str(value(row, "月份") or "").strip()
            if created_at is None or not sample_id or not hospital_name:
                continue

            cycle_id = str(value(row, "送检单编号") or "").strip()
            sample_name = str(value(row, "送检单样本名称") or "").strip()
            product_name = str(value(row, "产品英文简写") or "").strip()
            age_label = _normalize_age_label(value(row, "受检人年龄（人工处理）"))
            age = _parse_age(value(row, "受检人年龄（系统）", "受检人年龄"))

            records.append(
                DetailRecord(
                    created_at=created_at,
                    month_bucket=month_bucket or created_at.strftime("%Y-%m"),
                    cycle_id=cycle_id,
                    task_name=cycle_id,
                    sample_id=sample_id,
                    sample_name=sample_name,
                    sample_type=sample_type,
                    hospital_name=hospital_name,
                    hospital_code=hospital_name,
                    product_name=product_name,
                    description_cn=_normalize_result_label(value(row, "结果解释")),
                    description_en="",
                    qc_conclusion=_normalize_qc_conclusion(value(row, "data_QC_conclusion")),
                    ldpgta_qc_conclusion="",
                    age=age,
                    age_label=age_label,
                    mapd=None,
                    bincv=_to_float(value(row, "CV(1000K_bin_size)")),
                    cnvpq=None,
                    cnv_type=_normalize_text(value(row, "CNV检测结果")),
                    upd_arms=_normalize_text(value(row, "染色体位置")),
                    ldpgta_cnv_type=_normalize_text(value(row, "提示CNV")),
                    result_detail=_normalize_text(value(row, "结果说明")),
                    incidental_label=_normalize_binary_flag(value(row, "意外发现（人工处理）")),
                    aneuploidy_label=_normalize_aneuploidy_label(value(row, "异倍体结果（人工处理）")),
                )
            )
        return records

    @property
    def records(self) -> list[DetailRecord]:
        return self._records

    @property
    def eligible_records(self) -> list[DetailRecord]:
        return [record for record in self.records if record.is_embryo_sample]

    @property
    def snapshot_range(self) -> tuple[datetime, datetime]:
        eligible = self.eligible_records
        return min(item.created_at for item in eligible), max(item.created_at for item in eligible)

    @property
    def stat_month_range(self) -> tuple[str, str]:
        buckets = sorted({item.month_bucket for item in self.eligible_records if item.month_bucket})
        if not buckets:
            fallback_start, fallback_end = self.snapshot_range
            return fallback_start.strftime("%Y-%m"), fallback_end.strftime("%Y-%m")
        return buckets[0], buckets[-1]

    def filter_records(
        self,
        *,
        hospital_id: str | None = None,
        year: int | None = None,
        quarter: int | None = None,
        month: int | None = None,
        day: int | None = None,
        end_year: int | None = None,
        end_month: int | None = None,
        start_day: date | None = None,
        end_day: date | None = None,
        age_range: str | None = None,
    ) -> list[DetailRecord]:
        filtered = self.eligible_records
        if hospital_id:
            filtered = [item for item in filtered if item.hospital_code == hospital_id]
        if start_day and end_day:
            filtered = [
                item for item in filtered if start_day <= item.created_at.date() <= end_day
            ]
        else:
            if year is not None:
                if end_year is not None and end_month is not None and month is not None:
                    filtered = [
                        item
                        for item in filtered
                        if _in_month_range(
                            item=item,
                            start_year=year,
                            start_month=month,
                            end_year=end_year,
                            end_month=end_month,
                        )
                    ]
                else:
                    filtered = [item for item in filtered if _bucket_year(item.month_bucket, item.created_at) == year]
            if quarter is not None:
                filtered = [item for item in filtered if _bucket_quarter(item.month_bucket, item.created_at) == quarter]
            if month is not None and end_month is None:
                filtered = [item for item in filtered if _bucket_month(item.month_bucket, item.created_at) == month]
            if day is not None:
                filtered = [item for item in filtered if item.created_at.day == day]
        if age_range:
            filtered = [item for item in filtered if _match_age_range(item.age, item.age_label, age_range)]
        return filtered

    @property
    def hospitals(self) -> list[dict[str, str | int]]:
        counts = Counter(item.hospital_name for item in self.eligible_records)
        ordered = sorted(counts.items(), key=lambda item: (-item[1], item[0]))
        return [
            {
                "hospital_id": hospital_name,
                "hospital_name": hospital_name,
                "sample_count": sample_count,
            }
            for hospital_name, sample_count in ordered
        ]


@lru_cache(maxsize=1)
def get_pgta_dataset() -> PGTADetailDataset:
    return PGTADetailDataset(settings.pgta_detail_file)


def _bucket_year(month_bucket: str, fallback: datetime) -> int:
    if parsed := _parse_month_bucket(month_bucket):
        return parsed.year
    return fallback.year


def _bucket_month(month_bucket: str, fallback: datetime) -> int:
    if parsed := _parse_month_bucket(month_bucket):
        return parsed.month
    return fallback.month


def _bucket_quarter(month_bucket: str, fallback: datetime) -> int:
    month = _bucket_month(month_bucket, fallback)
    return ((month - 1) // 3) + 1


def _parse_month_bucket(value: str) -> datetime | None:
    text = (value or "").strip()
    if not text:
        return None
    for fmt in ("%Y-%m", "%Y/%m", "%Y年%m月"):
        try:
            return datetime.strptime(text, fmt)
        except ValueError:
            continue
    return None


def _in_month_range(
    *,
    item: DetailRecord,
    start_year: int,
    start_month: int,
    end_year: int,
    end_month: int,
) -> bool:
    year, month = _bucket_year(item.month_bucket, item.created_at), _bucket_month(item.month_bucket, item.created_at)
    current = year * 12 + month
    start = start_year * 12 + start_month
    end = end_year * 12 + end_month
    return start <= current <= end


def _match_age_range(age: int | None, age_label: str, age_range: str) -> bool:
    if age_range == "missing":
        return age is None or age_label == "未填写"
    if age_range == "bucket:35-38":
        return age_label == "35-38岁"
    if age_range == "bucket:gt38":
        return age_label == "＞38岁"
    if age_range == "bucket:lt35":
        return age_label == "＜35岁"
    if age_range.startswith("gt:"):
        threshold = int(age_range.split(":", 1)[1])
        if threshold == 35 and age_label in {"35-38岁", "＞38岁"}:
            return True
        if age is not None:
            return age > threshold
        return False
    if age_range.startswith("gte:"):
        threshold = int(age_range.split(":", 1)[1])
        if threshold == 35 and age_label in {"35-38岁", "＞38岁"}:
            return True
        if age is not None:
            return age >= threshold
        return False
    if age_range.startswith("lt:"):
        threshold = int(age_range.split(":", 1)[1])
        if age is not None:
            return age < threshold
        return age_label == "＜35岁" if threshold == 35 else False
    if age_range.startswith("lte:"):
        threshold = int(age_range.split(":", 1)[1])
        if age is not None:
            return age <= threshold
        return age_label in {"＜35岁", "35-38岁"} if threshold in {37, 38} else False
    if age_range.startswith("between:"):
        start, end = age_range.split(":", 1)[1].split(",", 1)
        start_i = int(start)
        end_i = int(end)
        if start_i == 35 and end_i in {37, 38} and age_label == "35-38岁":
            return True
        if age is not None:
            return start_i <= age <= end_i
        return False
    if age is None:
        return False
    return True
