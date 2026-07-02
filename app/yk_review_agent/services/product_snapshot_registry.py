from __future__ import annotations

from dataclasses import dataclass
from datetime import date, datetime
from functools import lru_cache
from pathlib import Path

from openpyxl import load_workbook

from yk_review_agent.core.config import settings


CANONICAL_FIELD_ALIASES: dict[str, tuple[str, ...]] = {
    "hospital": ("送检单位名称",),
    "stat_month": ("月份",),
    "review_time": ("报告审核时间", "项目日期"),
    "cycle_id": ("送检单编号",),
    "sample_id": ("样本编号",),
    "sample_name": ("送检单样本名称",),
    "sample_type": ("样本类型",),
    "product_code": ("产品英文简写",),
    "result_label": ("结果解释",),
    "qc_result": ("data_QC_conclusion", "QC_conclusion"),
    "female_age": ("受检人年龄（人工处理）", "受检人年龄（系统）", "受检人年龄"),
    "bin_cv": ("CV(1000K_bin_size)",),
    "cnv_result": ("CNV检测结果",),
    "cnv_hint": ("提示CNV",),
    "chromosome_location": ("染色体位置",),
    "result_detail": ("结果说明",),
    "clinical_indicator": ("临床指征（人工处理）", "临床指征（系统）", "临床指征"),
    "incidental_label": ("意外发现（人工处理）", "有无意外发现", "意外发现人工处理）"),
    "aneuploidy_label": ("异倍体结果（人工处理）", "异倍体分析结果（对应亲本污染分析字段）"),
    "gene_name": ("基因",),
    "disease_name": ("检测疾病名称",),
    "variant_result": ("变异携带结果",),
    "family_type": ("家系类型", "家系代码"),
    "marecs_stage2": ("是否转MaReCs第二阶段",),
    "marecs_family_type": ("二阶段家系类型",),
    "next_step_screening": ("是否进入下一步易位筛查",),
}


@dataclass(frozen=True)
class SnapshotSourceConfig:
    product_code: str
    product_label: str
    file_path: str
    sheet_name: str
    execution_status: str
    notes: tuple[str, ...]


@dataclass(frozen=True)
class SnapshotSourceProfile:
    product_code: str
    product_label: str
    data_source: str
    sheet_name: str
    row_count: int
    cycle_count: int
    snapshot_start: str | None
    snapshot_end: str | None
    semantic_fields: tuple[str, ...]
    execution_status: str
    notes: tuple[str, ...]


SOURCE_CONFIGS: tuple[SnapshotSourceConfig, ...] = (
    SnapshotSourceConfig(
        product_code="PGT-A",
        product_label="PGT-A",
        file_path=settings.pgta_detail_file,
        sheet_name="2025年-数据",
        execution_status="executable",
        notes=(
            "当前真实执行主链路使用这份快照。",
            "包含人工处理后的年龄、临床指征、意外发现和异倍体结果字段。",
        ),
    ),
    SnapshotSourceConfig(
        product_code="PGT-AH",
        product_label="PGT-AH",
        file_path=settings.pgtah_snapshot_file,
        sheet_name="2025源数据",
        execution_status="metadata_only",
        notes=(
            "快照已接入，可用于能力边界说明。",
            "当前尚未接入执行函数层。",
        ),
    ),
    SnapshotSourceConfig(
        product_code="PGT-SR",
        product_label="PGT-SR（含 MaReCs 相关字段）",
        file_path=settings.pgtsr_snapshot_file,
        sheet_name="2025年-数据",
        execution_status="metadata_only",
        notes=(
            "快照包含 MaReCs 第二阶段相关字段。",
            "当前尚未接入真实统计执行。",
        ),
    ),
    SnapshotSourceConfig(
        product_code="PGT-M",
        product_label="PGT-M",
        file_path=settings.pgtm_snapshot_file,
        sheet_name="原始数据",
        execution_status="metadata_only",
        notes=(
            "快照包含基因、疾病、家系类型和携带状态等专有字段。",
            "工作簿存在 Excel 格式污染，需要按有效记录扫描。",
        ),
    ),
)


class WorkbookSnapshotAdapter:
    def __init__(self, config: SnapshotSourceConfig) -> None:
        self.config = config
        self.file_path = Path(config.file_path)
        self.sheet_name = config.sheet_name

    @property
    def workbook_name(self) -> str:
        return self.file_path.name

    @lru_cache(maxsize=1)
    def profile(self) -> SnapshotSourceProfile:
        workbook = load_workbook(self.file_path, read_only=True, data_only=True)
        sheet_name = self._resolve_sheet_name(workbook.sheetnames)
        sheet = workbook[sheet_name]
        header = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
        index = {
            str(value).strip(): position
            for position, value in enumerate(header)
            if value is not None and str(value).strip()
        }

        sample_idx = self._first_index(index, "sample_id")
        cycle_idx = self._first_index(index, "cycle_id")
        time_idx = self._first_index(index, "review_time")
        sample_type_idx = self._first_index(index, "sample_type")

        row_count = 0
        cycle_ids: set[str] = set()
        semantic_fields: set[str] = set()
        snapshot_start: datetime | None = None
        snapshot_end: datetime | None = None
        empty_streak = 0

        for row in sheet.iter_rows(min_row=2, values_only=True):
            if not any(value is not None and str(value).strip() for value in row):
                empty_streak += 1
                if self.config.product_code == "PGT-M" and empty_streak >= 5000:
                    break
                continue

            empty_streak = 0
            sample_id = self._value(row, sample_idx)
            sample_type = str(self._value(row, sample_type_idx) or "")
            if not sample_id or "对照" in sample_type or "极体" in sample_type:
                continue

            row_count += 1
            cycle_id = self._value(row, cycle_idx)
            if cycle_id:
                cycle_ids.add(str(cycle_id).strip())

            dt = _parse_datetime(self._value(row, time_idx))
            if dt is not None:
                if snapshot_start is None or dt < snapshot_start:
                    snapshot_start = dt
                if snapshot_end is None or dt > snapshot_end:
                    snapshot_end = dt

        for canonical_field, aliases in CANONICAL_FIELD_ALIASES.items():
            if any(alias in index for alias in aliases):
                semantic_fields.add(canonical_field)

        return SnapshotSourceProfile(
            product_code=self.config.product_code,
            product_label=self.config.product_label,
            data_source=self.workbook_name,
            sheet_name=sheet_name,
            row_count=row_count,
            cycle_count=len(cycle_ids),
            snapshot_start=snapshot_start.isoformat() if snapshot_start else None,
            snapshot_end=snapshot_end.isoformat() if snapshot_end else None,
            semantic_fields=tuple(sorted(semantic_fields)),
            execution_status=self.config.execution_status,
            notes=self.config.notes,
        )

    def _first_index(self, index: dict[str, int], canonical_field: str) -> int | None:
        for alias in CANONICAL_FIELD_ALIASES[canonical_field]:
            if alias in index:
                return index[alias]
        return None

    def _resolve_sheet_name(self, available_names: list[str]) -> str:
        if self.sheet_name in available_names:
            return self.sheet_name

        fallback_candidates = (
            "2025年-数据",
            "2025源数据",
            "原始数据",
        )
        for candidate in fallback_candidates:
            if candidate in available_names:
                return candidate
        return available_names[0]

    def _value(self, row: tuple[object, ...], position: int | None) -> object | None:
        if position is None or position >= len(row):
            return None
        value = row[position]
        if value is None:
            return None
        if isinstance(value, str) and not value.strip():
            return None
        return value


class ProductSnapshotRegistry:
    def __init__(self) -> None:
        self._adapters = {
            config.product_code: WorkbookSnapshotAdapter(config)
            for config in SOURCE_CONFIGS
        }

    def list_profiles(self) -> list[SnapshotSourceProfile]:
        return [self._adapters[config.product_code].profile() for config in SOURCE_CONFIGS]

    def get_profile(self, product_code: str) -> SnapshotSourceProfile | None:
        adapter = self._adapters.get(product_code)
        return adapter.profile() if adapter else None

    def active_capabilities(self, product_code: str) -> frozenset[str]:
        profile = self.get_profile(product_code)
        if profile is None:
            return frozenset()
        return frozenset(profile.semantic_fields)

    def executable_products(self) -> tuple[str, ...]:
        return tuple(
            config.product_code
            for config in SOURCE_CONFIGS
            if config.execution_status == "executable"
        )


def _parse_datetime(value: object) -> datetime | None:
    if isinstance(value, datetime):
        return value
    if isinstance(value, date):
        return datetime.combine(value, datetime.min.time())
    if value is None:
        return None
    text = str(value).strip()
    if not text:
        return None
    for fmt in (
        "%Y-%m-%d %H:%M:%S",
        "%Y-%m-%d",
        "%Y/%m/%d %H:%M:%S",
        "%Y/%m/%d",
        "%Y年%m月%d日",
    ):
        try:
            return datetime.strptime(text, fmt)
        except ValueError:
            continue
    return None


@lru_cache(maxsize=1)
def get_snapshot_registry() -> ProductSnapshotRegistry:
    return ProductSnapshotRegistry()
