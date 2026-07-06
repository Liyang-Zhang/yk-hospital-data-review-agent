from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook

from yk_review_agent.core.config import settings
from yk_review_agent.services.pgta_record_source import clear_pgta_record_source_cache
from yk_review_agent.services.pgta_sqlite import (
    PGTAExcelImporter,
    PGTASQLiteRepository,
    PGTASourceConfig,
    inspect_pgta_snapshot_db,
)
from yk_review_agent.services.query_service import query_service


BASE_FILTERS = {
    "hospital_id": "中国人民解放军医院301医院",
    "hospital_name": "中国人民解放军医院301医院",
}


def _build_temp_db(tmp_path: Path, workbook_path: Path, *, sheet_name: str) -> str:
    database_url = f"sqlite+pysqlite:///{(tmp_path / 'snapshot.db').resolve()}"
    PGTAExcelImporter(
        database_url,
        source_configs=[
            PGTASourceConfig(
                snapshot_year=2026,
                snapshot_half="test",
                file_path=str(workbook_path),
                sheet_name=sheet_name,
                column_map={},
            )
        ],
    ).rebuild()
    return database_url


def test_header_normalization_handles_newlines_in_manual_columns(tmp_path: Path) -> None:
    workbook_path = tmp_path / "pgta-2026.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "2026年-数据"
    ws.append(
        [
            "月份",
            "报告审核时间",
            "送检单位名称",
            "产品英文简写",
            "送检单编号",
            "样本编号",
            "送检单样本名称",
            "样本类型",
            "受检人年龄",
            "受检人年龄\n（人工处理）",
            "结果解释",
            "data_QC_conclusion",
            "CNV检测结果",
            "提示CNV",
            "染色体位置",
            "结果说明",
            "意外发现\n（人工处理）",
            "异倍体结果\n（人工处理）",
            "CV(1000K_bin_size)",
        ]
    )
    ws.append(
        [
            "2026-01",
            "2026-01-15 12:00:00",
            "测试医院",
            "PGT-A",
            "cycle-1",
            "sample-1",
            "sample-1",
            "胚胎",
            36,
            "35-38岁",
            "未见异常",
            "PASS",
            "无",
            "del(1)(p36.33)(~1.20Mb)",
            "Xp22.33",
            "该区域包含已知综合征",
            "有",
            "未见异常",
            0.12,
        ]
    )
    wb.save(workbook_path)

    database_url = _build_temp_db(tmp_path, workbook_path, sheet_name="2026年-数据")
    actual_repo = PGTASQLiteRepository(database_url)
    record = actual_repo.eligible_records[0]

    assert record.age_label == "35-38岁"
    assert record.incidental_label == "有"
    assert record.aneuploidy_label == "未见异常"


def test_short_year_month_bucket_is_normalized(tmp_path: Path) -> None:
    workbook_path = tmp_path / "pgta-2024.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "源数据"
    ws.append(
        [
            "月份",
            "报告审核时间",
            "送检单位名称",
            "产品英文简写",
            "送检单编号",
            "样本编号",
            "送检单样本名称",
            "样本类型",
            "受检人年龄",
            "结果解释",
            "data_QC_conclusion",
            "CNV检测结果",
            "提示CNV",
            "染色体位置",
            "结果说明",
            "CV(1000K_bin_size)",
        ]
    )
    ws.append(
        [
            "2411",
            "2024-11-06 00:00:00",
            "测试医院",
            "PGT-A",
            "cycle-1",
            "sample-1",
            "sample-1",
            "胚胎",
            36,
            "未见异常",
            "PASS",
            "无",
            "hint",
            "Xp22.33",
            "detail",
            0.12,
        ]
    )
    wb.save(workbook_path)

    database_url = f"sqlite+pysqlite:///{(tmp_path / 'short-month.db').resolve()}"
    PGTAExcelImporter(
        database_url,
        source_configs=[
            PGTASourceConfig(
                snapshot_year=2024,
                snapshot_half="full",
                file_path=str(workbook_path),
                sheet_name="源数据",
                column_map={},
            )
        ],
    ).rebuild()

    actual_repo = PGTASQLiteRepository(database_url)
    record = actual_repo.eligible_records[0]
    assert record.month_bucket == "2024-11"
    summary = inspect_pgta_snapshot_db(database_url)
    assert summary["snapshot_start"] == "2024-11"
    assert summary["snapshot_end"] == "2024-11"


def test_core_field_value_normalization_handles_obvious_aliases(tmp_path: Path) -> None:
    workbook_path = tmp_path / "pgta-core-normalization.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "源数据"
    ws.append(
        [
            "月份",
            "报告审核时间",
            "送检单位名称",
            "产品英文简写",
            "送检单编号",
            "样本编号",
            "送检单样本名称",
            "样本类型",
            "受检人年龄",
            "结果解释",
            "data_QC_conclusion",
            "CNV检测结果",
            "提示CNV",
            "染色体位置",
            "结果说明",
            "是否有意外发现",
            "异倍体分析",
            "CV(1000K_bin_size)",
        ]
    )
    ws.append(
        [
            "2408",
            "2024-08-11 00:00:00",
            "测试医院",
            "PGT-A",
            "cycle-1",
            "sample-1",
            "sample-1",
            "囊胚滋养层细胞",
            36,
            "未见染色体拷贝数异常",
            "XX",
            "无",
            "hint",
            "Xp22.33",
            "detail",
            "是",
            "/",
            0.12,
        ]
    )
    wb.save(workbook_path)

    database_url = f"sqlite+pysqlite:///{(tmp_path / 'core-normalization.db').resolve()}"
    PGTAExcelImporter(
        database_url,
        source_configs=[
            PGTASourceConfig(
                snapshot_year=2024,
                snapshot_half="full",
                file_path=str(workbook_path),
                sheet_name="源数据",
                column_map={},
            )
        ],
    ).rebuild()

    record = PGTASQLiteRepository(database_url).eligible_records[0]
    assert record.description_cn == "未见异常"
    assert record.qc_conclusion == ""
    assert record.incidental_label == "有"
    assert record.aneuploidy_label == ""


def test_sqlite_backend_matches_excel_backend_for_core_metric() -> None:
    original_backend = settings.snapshot_backend
    try:
        settings.snapshot_backend = "sqlite"
        clear_pgta_record_source_cache()
        sqlite_result = query_service.run(
            "pgta_euploid_rate",
            {**BASE_FILTERS, "time_range": "2025年", "breakdown": "overall", "focus": "summary"},
        )

        settings.snapshot_backend = "excel"
        clear_pgta_record_source_cache()
        excel_result = query_service.run(
            "pgta_euploid_rate",
            {**BASE_FILTERS, "time_range": "2025年", "breakdown": "overall", "focus": "summary"},
        )
    finally:
        settings.snapshot_backend = original_backend
        clear_pgta_record_source_cache()

    assert sqlite_result["table"]["rows"] == excel_result["table"]["rows"]
    assert sqlite_result["summary"] == excel_result["summary"]


def test_snapshot_db_inspection_returns_expected_shape() -> None:
    summary = inspect_pgta_snapshot_db(settings.snapshot_db_url)
    assert summary["table_name"] == "pgta_snapshot_raw"
    assert summary["row_count"] > 0
    assert summary["hospital_count"] > 0
    assert summary["snapshot_start"]
    assert summary["snapshot_end"]
